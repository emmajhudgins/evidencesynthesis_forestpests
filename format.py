from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from xlsxwriter.utility import xl_rowcol_to_cell
import xlsxwriter
import pandas as pd
import numpy as np
import re

FILENAMES = [
    './sheets/review/review_scopus_screen.xlsx',
    './sheets/review/webofscience_review_screen.xlsx',
    './sheets/evidence/scopus_screen.xlsx',
    './sheets/evidence/webofscience_screen.xlsx',
    './sheets/review/review_unique_wos.xlsx',
    './sheets/evidence/unique_wos.xlsx']

DATAFRAMES = [pd.read_excel(f) for f in FILENAMES]

PEST_INCLUDED = pd.read_csv('./sheets/pest_traits.csv')
PEST_INCLUDED = np.concatenate([PEST_INCLUDED['LATIN_NAME'].to_numpy(), PEST_INCLUDED['COMMON_NAM'].to_numpy()], axis=0)
PEST_INCLUDED = [name.lower() for name in PEST_INCLUDED]

PEST_EXCLUDED = ['Lyme disease', 'Borrelia Burgdorferi',
                 'Mountain pine beetle', 'Dendroctonus ponderosae',
                 'Western Spruce budworm', 'Choristoneura occidentalis',
                 'Eastern Spruce budworm', 'Choristoneura fumiferana']
PEST_EXCLUDED = [name.lower() for name in PEST_EXCLUDED]

KEYWORDS = [
    'forest', 'invasi', 'non-native', 'exotic', 'alien', 'pest', 'pathogen',
    'insect', 'manag', 'policy', 'control', 'respon', 'prevent', 'surve',
    'eradict', 'detect', 'canada', 'united states', 'north america']

DFs = [pd.read_excel(f) for f in FILENAMES]

COL_NAMES = [['Title', 'Abstract', 'Author Keywords', 'Index Keywords'],
             ['Title', 'Keywords', 'Keywords Plus', 'Abstract'],
             ['Title', 'Abstract', 'Author Keywords', 'Index Keywords'],
             ['Title', 'Keywords', 'Keywords plus', 'Abstract']]

string = DFs[0]['Abstract'][0]


def get_indices(s, label, search_terms):
    indices = [[(m.start(), m.start() + len(w)) for m in re.finditer(w, s.lower())] for w in search_terms]
    return [(i, j, label) for sublist in indices for (i, j) in sublist]


def string_partition(lst, length):
    if len(lst) == 0:
        return []
    all_indices = []
    if lst[0][0] > 0:
        all_indices.append((0, lst[0][0], 'normal'))
    for i in range(len(lst) - 1):
        left = lst[i][1]
        right = lst[i + 1][0]
        all_indices.append(lst[i])
        if right > left:
            all_indices.append((left, right, 'normal'))
    all_indices.append(lst[len(lst) - 1])
    last = lst[len(lst) - 1][1]
    if last < length:
        all_indices.append((last, length, 'normal'))
    return all_indices


def highlight_keywords(df, col_name, path):
    col = df[col_name].to_numpy()

    workbook = xlsxwriter.Workbook(f'{path}/{col_name}.xlsx')
    worksheet = workbook.add_worksheet()

    # normal format
    normal_format = workbook.add_format({
        'color': 'black'
    })

    # format the keywords
    keyword_format = workbook.add_format({
        'color': 'red',
        'bold': True,
        'font_size': 20
    })

    # format the species to be included
    include_format = workbook.add_format({
        'color': 'blue',
        'bold': True,
        'font_size': 20
    })

    # format the species to be excluded
    exclude_format = workbook.add_format({
        'bold': True,
        'font_size': 20
    })
    exclude_format.set_font_strikeout()

    formats = {'normal': normal_format,
               'include': include_format,
               'exclude': exclude_format,
               'keyword': keyword_format}

    cell = xl_rowcol_to_cell(0, 0)
    worksheet.write(cell, col_name + ' Highlighted')

    for i, sentence in enumerate(col):
        cell = xl_rowcol_to_cell(i+1, 0)
        if type(sentence) != str:
            print('invalid sentence: ', sentence)
            sentence = ""
        include_indices = get_indices(sentence, 'include', PEST_INCLUDED)
        exclude_indices = get_indices(sentence, 'exclude', PEST_EXCLUDED)
        keyword_indices = get_indices(sentence, 'keyword', KEYWORDS)
        indices = include_indices + exclude_indices + keyword_indices
        sorted_indices = sorted(indices)
        all_indices = string_partition(sorted_indices, len(sentence))
        if all_indices:
            formatted = []
            for (i, j, label) in all_indices:
                format = formats[label]
                token = sentence[i:j]
                formatted.append(format)
                formatted.append(token)
                worksheet.write_rich_string(cell, *formatted)
        else:
            worksheet.write(cell, sentence)

    workbook.close()



for col in COL_NAMES[1]:
    highlight_keywords(DFs[5], col, './sheets/evidence/formatted/unique')


def highlight_sentence(df, col_name, path, keywords):
    col = df[col_name].to_numpy()

    workbook = xlsxwriter.Workbook(f'{path}/{col_name}.xlsx')
    worksheet = workbook.add_worksheet()

    highlight = workbook.add_format({
        'color': 'red',
        'bold': True,
        'font_size': 20
    })

    cell = xl_rowcol_to_cell(0, 0)
    worksheet.write(cell, 'Sentence Highlighted')
    for i, e in enumerate(col):
        sentences = e.split('.')
        formatted = []
        for s in sentences:
            for k in keywords:
                if k in s.lower():
                    formatted.append(highlight)
                    break
            if s != '':
                formatted.append(s)
                formatted.append('.')

        cell = xl_rowcol_to_cell(i+1, 0)
        worksheet.write_rich_string(cell, *formatted)

    workbook.close()

highlight_sentence(DFs[5], 'Abstract', './sheets/evidence/sentence/unique', KEYWORDS)
